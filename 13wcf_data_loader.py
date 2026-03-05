"""
13WCF Data Loader - Python Version
===================================
Loads data from Settlement, Supplier, and Customer files into the 13WCF Model

Usage:
    python 13wcf_data_loader.py

The script will prompt you for file paths.
"""

import os
import sys
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter


def log(msg):
    """Print log message with timestamp"""
    timestamp = datetime.now().strftime("%H:%M:%S")
    print(f"[{timestamp}] {msg}")


def get_file_path(prompt):
    """Prompt user for file path and validate it exists"""
    while True:
        path = input(f"\n{prompt}\nPath: ").strip().strip('"').strip("'")
        if os.path.exists(path):
            return path
        print(f"❌ File not found: {path}")
        print("Please try again or press Ctrl+C to exit")


def adjust_formula(formula, offset):
    """Adjust row references in Excel formula by offset"""
    if not formula:
        return ""

    import re
    def replace_ref(match):
        col = match.group(1)
        abs_marker = match.group(2)
        row = match.group(3)

        # Don't adjust absolute row references
        if abs_marker == '$':
            return match.group(0)

        new_row = int(row) + offset
        return f"{col}{new_row}"

    return re.sub(r'(\$?[A-Z]{1,3})(\$?)(\d+)', replace_ref, formula)


def process_settlement_run(wb, settlement_path):
    """Process Settlement Details → OPEX - Weekly Settlement Run"""
    log("Processing: OPEX - Weekly Settlement Run...")

    # Open settlement file
    wb_settle = openpyxl.load_workbook(settlement_path, data_only=True, read_only=True)
    ws_settle = wb_settle.worksheets[0]

    # Get target worksheet
    ws_target = wb.worksheets[wb.sheetnames.index('OPEX - Weekly Settlement Run')]

    # Save formula templates from row 9 (columns AT-BD, which are 46-56)
    formula_templates = {}
    for col in range(46, 57):
        cell = ws_target.cell(9, col)
        if cell.value and str(cell.value).startswith('='):
            formula_templates[col] = str(cell.value)

    # Clear old data from row 9 onwards (columns A-AR = 1-44, AT-BD = 46-56)
    max_row = ws_target.max_row
    for row in range(9, max_row + 1):
        for col in range(1, 45):
            ws_target.cell(row, col).value = None
        for col in range(46, 57):
            ws_target.cell(row, col).value = None

    # Paste settlement data (skip first 2 header rows)
    rows_pasted = 0
    for row_idx, row in enumerate(ws_settle.iter_rows(min_row=3, values_only=True)):
        # Check if row has any data
        if not any(row[:44]):
            continue

        dest_row = 9 + rows_pasted

        # Copy data columns A-AR (1-44)
        for col in range(1, 45):
            if col - 1 < len(row):
                ws_target.cell(dest_row, col).value = row[col - 1]

        # Copy adjusted formulas AT-BD (46-56)
        for col, formula in formula_templates.items():
            adjusted = adjust_formula(formula, rows_pasted)
            ws_target.cell(dest_row, col).value = adjusted

        rows_pasted += 1
        if rows_pasted % 500 == 0:
            log(f"  Settlement: {rows_pasted} rows processed...")

    wb_settle.close()
    log(f"✓ Settlement Run: {rows_pasted} rows loaded into rows 9-{8 + rows_pasted}")
    return rows_pasted


def process_supplier_invoices(wb, supplier_path):
    """Process Supplier Invoices → OPEX - AP"""
    log("Processing: OPEX - AP...")

    # Open supplier file
    wb_supplier = openpyxl.load_workbook(supplier_path, data_only=True, read_only=True)
    ws_supplier = wb_supplier.worksheets[0]

    # Get target worksheet
    ws_target = wb.worksheets[wb.sheetnames.index('OPEX - AP')]

    # Save formula templates from row 13 (columns AO-BD, which are 41-56)
    formula_templates = {}
    for col in range(41, 57):
        cell = ws_target.cell(13, col)
        if cell.value and str(cell.value).startswith('='):
            formula_templates[col] = str(cell.value)

    # Clear old data from row 13 onwards (columns A-AM = 1-39, AO-BD = 41-56)
    max_row = ws_target.max_row
    for row in range(13, max_row + 1):
        for col in range(1, 40):
            ws_target.cell(row, col).value = None
        for col in range(41, 57):
            ws_target.cell(row, col).value = None

    # Paste supplier data (skip first 2 header rows)
    rows_pasted = 0
    for row_idx, row in enumerate(ws_supplier.iter_rows(min_row=3, values_only=True)):
        # Check if row has any data
        if not any(row[:39]):
            continue

        dest_row = 13 + rows_pasted

        # Copy data columns A-AM (1-39)
        for col in range(1, 40):
            if col - 1 < len(row):
                ws_target.cell(dest_row, col).value = row[col - 1]

        # Copy adjusted formulas AO-BD (41-56)
        for col, formula in formula_templates.items():
            adjusted = adjust_formula(formula, rows_pasted)
            ws_target.cell(dest_row, col).value = adjusted

        rows_pasted += 1
        if rows_pasted % 500 == 0:
            log(f"  Supplier: {rows_pasted} rows processed...")

    wb_supplier.close()
    log(f"✓ Supplier Invoices: {rows_pasted} rows loaded into rows 13-{12 + rows_pasted}")
    return rows_pasted


def process_customer_invoices(wb, customer_path):
    """Process Customer Invoices → AR - Customer Invoice Detail"""
    log("Processing: AR - Customer Invoice Detail...")

    # Open customer file
    wb_customer = openpyxl.load_workbook(customer_path, data_only=True, read_only=True)
    ws_customer = wb_customer.worksheets[0]

    # Get target worksheet
    ws_target = wb.worksheets[wb.sheetnames.index('AR - Customer Invoice Detail')]

    # Save formula templates from row 8 (columns R-Y, which are 18-25)
    formula_templates = {}
    for col in range(18, 26):
        cell = ws_target.cell(8, col)
        if cell.value and str(cell.value).startswith('='):
            formula_templates[col] = str(cell.value)

    # Clear old data from row 8 onwards (columns A-Q = 1-17, R-Y = 18-25)
    max_row = ws_target.max_row
    for row in range(8, max_row + 1):
        for col in range(1, 18):
            ws_target.cell(row, col).value = None
        for col in range(18, 26):
            ws_target.cell(row, col).value = None

    # Paste customer data (skip first 2 header rows)
    rows_pasted = 0
    for row_idx, row in enumerate(ws_customer.iter_rows(min_row=3, values_only=True)):
        # Check if row has any data
        if not any(row[:17]):
            continue

        dest_row = 8 + rows_pasted

        # Copy data columns A-Q (1-17)
        for col in range(1, 18):
            if col - 1 < len(row):
                ws_target.cell(dest_row, col).value = row[col - 1]

        # Copy adjusted formulas R-Y (18-25)
        for col, formula in formula_templates.items():
            adjusted = adjust_formula(formula, rows_pasted)
            ws_target.cell(dest_row, col).value = adjusted

        rows_pasted += 1
        if rows_pasted % 500 == 0:
            log(f"  Customer: {rows_pasted} rows processed...")

    wb_customer.close()
    log(f"✓ Customer Invoices: {rows_pasted} rows loaded into rows 8-{7 + rows_pasted}")
    return rows_pasted


def main():
    print("\n" + "="*60)
    print("13WCF Data Loader - Python Version")
    print("="*60)

    # Get file paths
    print("\nPlease provide the following files:")
    print("(You can drag & drop files into this window)\n")

    source_path = get_file_path("1. 13WCF Model (Bullish - 13WCF - DRAFT .xlsx)")
    settlement_path = get_file_path("2. Settlement Details (Settlement_details.xlsx)")
    supplier_path = get_file_path("3. Supplier Invoices (Find_Supplier_Invoices_-_Devi.xlsx)")
    customer_path = get_file_path("4. Customer Invoices (Customer_Invoice_Details.xlsx)")

    print("\n" + "="*60)
    log("Starting data processing...")
    print("="*60 + "\n")

    try:
        # Load source workbook
        log("Loading 13WCF Model...")
        wb = openpyxl.load_workbook(source_path)
        log(f"  Loaded workbook with {len(wb.sheetnames)} sheets")

        # Process each input file
        rows_sr = process_settlement_run(wb, settlement_path)
        rows_ap = process_supplier_invoices(wb, supplier_path)
        rows_ar = process_customer_invoices(wb, customer_path)

        # Save output file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = f"Bullish_13WCF_Updated_{timestamp}.xlsx"

        log("Saving updated workbook...")
        wb.save(output_path)
        wb.close()

        # Summary
        print("\n" + "="*60)
        print("✓ PROCESSING COMPLETE!")
        print("="*60)
        print(f"\nSummary:")
        print(f"  Settlement Run:      {rows_sr:,} rows loaded")
        print(f"  Supplier Invoices:   {rows_ap:,} rows loaded")
        print(f"  Customer Invoices:   {rows_ar:,} rows loaded")
        print(f"\nOutput file: {output_path}")
        print(f"Location: {os.path.abspath(output_path)}")
        print("\n" + "="*60)

    except KeyError as e:
        print(f"\n❌ ERROR: Sheet not found in source file")
        print(f"   Missing sheet: {str(e)}")
        print(f"\n   Available sheets: {', '.join(wb.sheetnames)}")
        sys.exit(1)

    except Exception as e:
        print(f"\n❌ ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n❌ Cancelled by user")
        sys.exit(0)
